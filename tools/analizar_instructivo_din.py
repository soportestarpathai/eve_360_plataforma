#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analiza instructivo_din.xlsx contra el formulario DIN operacion_din.php
Agrega columna CUMPLE REGLAS y OBSERVACIONES
"""
import re
import sys

# Tags que el formulario DIN y din_xml_helper SÍ soportan (mapeo del instructivo)
TAGS_SOPORTADOS = {
    'archivo', 'informe', 'mes_reportado',
    'sujeto_obligado', 'clave_entidad_colegiada', 'clave_sujeto_obligado', 'clave_actividad', 'exento',
    'aviso', 'referencia_aviso', 'modificatorio', 'folio_modificacion', 'descripcion_modificacion',
    'prioridad', 'alerta', 'tipo_alerta', 'descripcion_alerta',
    'detalle_operaciones', 'datos_operacion', 'tipo_operacion',
    'desarrollos_inmobiliarios', 'datos_desarrollo', 'objeto_aviso_anterior', 'modificacion',
    'entidad_federativa', 'registro_licencia', 'caracteristicas_desarrollo',
    'codigo_postal', 'colonia', 'calle', 'tipo_desarrollo', 'descripcion_desarrollo',
    'monto_desarrollo', 'unidades_comercializadas', 'costo_unidad', 'otras_empresas',
    'aportaciones', 'fecha_aportacion', 'tipo_aportacion',
    'recursos_propios', 'datos_aportacion', 'aportacion_numerario',
    'instrumento_monetario', 'moneda', 'monto_aportacion', 'aportacion_fideicomiso', 'nombre_institucion',
    'aportacion_especie', 'descripcion_bien', 'monto_estimado',
    'socios', 'numero_socios', 'detalle_socios', 'datos_socio',
    'aportacion_anterior_socio', 'rfc_socio', 'tipo_persona_socio',
    'persona_fisica', 'nombre', 'apellido_paterno', 'apellido_materno', 'fecha_nacimiento',
    'rfc', 'curp', 'pais_nacionalidad', 'actividad_economica',
    'persona_moral', 'denominacion_razon', 'fecha_constitucion', 'giro_mercantil',
    'representante_apoderado',
    'fideicomiso', 'identificador_fideicomiso',
    'tipo_domicilio_socio', 'nacional', 'extranjero',
    'numero_exterior', 'numero_interior', 'pais', 'estado_provincia', 'ciudad_poblacion',
    'telefono', 'clave_pais', 'numero_telefono', 'correo_electronico',
    'detalle_aportaciones',
    'terceros', 'numero_terceros', 'detalle_terceros', 'datos_tercero',
    'tipo_tercero', 'descripcion_tercero', 'tipo_persona_tercero',
    'valor_inmueble_preventa',
    'prestamo_financiero', 'datos_prestamo', 'tipo_institucion', 'institucion',
    'tipo_credito', 'monto_prestamo', 'plazo_meses',
    'prestamo_no_financiero', 'detalle_acreedores', 'tipo_persona_acreedor',
    'financiamiento_bursatil', 'fecha_emision', 'monto_solicitado', 'monto_recibido',
}

# Tags contenedores (estructura existe, no son campos de formulario)
TAGS_ESTRUCTURA = {'archivo', 'informe', 'aviso', 'modificatorio', 'alerta', 'detalle_operaciones',
    'datos_operacion', 'desarrollos_inmobiliarios', 'datos_desarrollo', 'caracteristicas_desarrollo',
    'aportaciones', 'tipo_aportacion', 'recursos_propios', 'datos_aportacion', 'aportacion_numerario',
    'aportacion_especie', 'socios', 'detalle_socios', 'datos_socio', 'tipo_persona_socio',
    'persona_fisica', 'persona_moral', 'fideicomiso', 'representante_apoderado',
    'tipo_domicilio_socio', 'nacional', 'extranjero', 'telefono',
    'terceros', 'detalle_terceros', 'datos_tercero', 'tipo_persona_tercero',
    'prestamo_financiero', 'datos_prestamo', 'prestamo_no_financiero',
    'detalle_acreedores', 'tipo_persona_acreedor', 'financiamiento_bursatil', 'sujeto_obligado'}

def extraer_tag(etiqueta_celda):
    """Extrae el nombre del tag de <tag> o tag"""
    if not etiqueta_celda:
        return None
    s = str(etiqueta_celda).strip()
    m = re.match(r'<?([a-z_]+)>?', s, re.I)
    return m.group(1).lower() if m else None

# Campos condicionales según instructivo (obligatorio solo en ciertos casos)
CONDICIONALES = {'nombre_institucion': 'Requerido si aportacion_fideicomiso=SI', 'clave_entidad_colegiada': 'Opcional si no aplica entidad colegiada'}

def cumple_reglas(tag, obligatoriedad):
    """Determina si el tag cumple con las reglas del formulario"""
    if not tag:
        return 'NO', 'Tag no identificado'
    if tag in TAGS_SOPORTADOS or tag in TAGS_ESTRUCTURA:
        obs = CONDICIONALES.get(tag, '')
        return 'SÍ', obs
    return 'NO', 'Campo no implementado en formulario DIN'

def main():
    try:
        import openpyxl
    except ImportError:
        print("Instalando openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        import openpyxl

    import os
    ruta_entrada = r'c:\Users\oscar\Downloads\instructivo_din(1).xlsx'
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ruta_salida = os.path.join(base_dir, 'instructivo_din_analizado.xlsx')
    ruta_downloads = r'c:\Users\oscar\Downloads\instructivo_din_analizado.xlsx'

    if not os.path.exists(ruta_entrada):
        print(f"Error: No se encontró {ruta_entrada}")
        sys.exit(1)

    wb = openpyxl.load_workbook(ruta_entrada, data_only=False)
    ws = wb['DIN']

    # Encontrar columna ETIQUETA XML (índice 2, columna C)
    # Encabezados: NO., NOMBRE DEL CAMPO, ETIQUETA XML, OBLIGATORIEDAD, ...
    headers = [cell.value for cell in ws[1]]
    col_etiqueta = None
    for i, h in enumerate(headers):
        if h and 'ETIQUETA' in str(h).upper():
            col_etiqueta = i + 1  # 1-based
            break
    if col_etiqueta is None:
        col_etiqueta = 3  # asumir columna C

    col_obligatoriedad = 4  # columna D

    # Insertar columnas CUMPLE REGLAS y OBSERVACIONES después de CATALOGOS (col 9)
    ws.insert_cols(10, 2)  # insertar 2 columnas en J
    ws.cell(1, 10, 'CUMPLE REGLAS')
    ws.cell(1, 11, 'OBSERVACIONES')

    for row_idx in range(2, ws.max_row + 1):
        etiqueta_val = ws.cell(row_idx, col_etiqueta).value
        oblig_val = ws.cell(row_idx, col_obligatoriedad).value or ''
        tag = extraer_tag(etiqueta_val)
        cumple, obs = cumple_reglas(tag, oblig_val)
        ws.cell(row_idx, 10, cumple)
        ws.cell(row_idx, 11, obs)

    wb.save(ruta_salida)
    if os.path.exists(os.path.dirname(ruta_downloads)):
        wb.save(ruta_downloads)
        print(f"Archivo guardado: {ruta_downloads}")
    print(f"También guardado en proyecto: {ruta_salida}")
    print("Columnas agregadas: CUMPLE REGLAS, OBSERVACIONES")

if __name__ == '__main__':
    main()

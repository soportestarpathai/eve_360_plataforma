#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Actualiza la columna COMPLETADO del instructivo_tsc.xlsx según lo implementado
en operacion_tsc.php y config/tsc_xml_helper.php (XSD TSC).

Ejecutar: python tools/actualizar_completado_instructivo_tsc.py
"""
import os
import sys

def main():
    try:
        import openpyxl
    except ImportError:
        print("Instalando openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        import openpyxl

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ruta = os.path.join(base_dir, 'instructivo_tsc.xlsx')

    if not os.path.exists(ruta):
        print(f"No se encontró {ruta}")
        sys.exit(1)

    wb = openpyxl.load_workbook(ruta, data_only=False)
    ws_name = 'F.II-1 Tarj ServCred' if 'F.II-1 Tarj ServCred' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]

    # Columna COMPLETADO (col 10)
    col_completado = 10
    if ws.cell(1, col_completado).value != 'COMPLETADO':
        # Buscar columna por nombre
        for c in range(1, ws.max_column + 1):
            if (ws.cell(1, c).value or '').strip() == 'COMPLETADO':
                col_completado = c
                break

    # Mapeo: fila -> valor COMPLETADO
    # SÍ = en formulario y XML
    # Pendiente = no implementado (ej. dueño beneficiario)
    # Estructural = generado automáticamente
    completado_map = {
        1: None,  # Header
        2: 'Estructural',
        3: 'Estructural',
        4: 'SÍ',
        5: 'Estructural',
        6: 'SÍ',
        7: 'SÍ',
        8: 'SÍ',
        9: 'SÍ',
        10: 'Estructural',
        11: 'SÍ',
        12: 'SÍ',
        13: 'SÍ',
        14: 'SÍ',
        15: 'SÍ',
        16: 'Estructural',
        17: 'SÍ',
        18: 'SÍ',
        19: 'Estructural',
        20: 'Estructural',
        21: 'Estructural',
        22: 'SÍ',
        23: 'SÍ',
        24: 'SÍ',
        25: 'SÍ',
        26: 'SÍ',
        27: 'SÍ',
        28: 'SÍ',
        29: 'SÍ',
        30: 'Estructural',
        31: 'SÍ',
        32: 'SÍ',
        33: 'SÍ',
        34: 'SÍ',
        35: 'SÍ',
        36: 'Estructural',
        37: 'SÍ',
        38: 'SÍ',
        39: 'SÍ',
        40: 'SÍ',
        41: 'SÍ',
        42: 'SÍ',
        43: 'Estructural',
        44: 'SÍ',
        45: 'SÍ',
        46: 'SÍ',
        47: 'Estructural',
        48: 'SÍ',
        49: 'SÍ',
        50: 'SÍ',
        51: 'SÍ',
        52: 'SÍ',
        53: 'SÍ',
        54: 'Estructural',
        55: 'Estructural',
        56: 'SÍ',
        57: 'SÍ',
        58: 'SÍ',
        59: 'SÍ',
        60: 'SÍ',
        61: 'Estructural',
        62: 'SÍ',
        63: 'SÍ',
        64: 'SÍ',
        65: 'SÍ',
        66: 'SÍ',
        67: 'SÍ',
        68: 'SÍ',
        69: 'SÍ',
        70: 'Estructural',
        71: 'SÍ',
        72: 'SÍ',
        73: 'SÍ',
        # 3.6 Dueño beneficiario - SÍ (implementado)
        74: 'Estructural',
        75: 'Estructural',
        76: 'Estructural',
        77: 'SÍ',
        78: 'SÍ',
        79: 'SÍ',
        80: 'SÍ',
        81: 'SÍ',
        82: 'SÍ',
        83: 'SÍ',
        84: 'SÍ',
        85: 'SÍ',
        86: 'Estructural',
        87: 'SÍ',
        88: 'SÍ',
        89: 'SÍ',
        90: 'Estructural',
        91: 'SÍ',
        92: 'SÍ',
        93: 'Estructural',
        94: 'Estructural',
        95: 'SÍ',
        96: 'SÍ',
        97: 'SÍ',
        98: 'SÍ',
        99: 'SÍ',
    }

    updated = 0
    for row, valor in completado_map.items():
        if row <= ws.max_row and valor:
            ws.cell(row, col_completado, valor)
            updated += 1

    wb.save(ruta)
    print(f"OK Instructivo actualizado: {ruta}")
    print(f"OK Columna COMPLETADO: {updated} celdas actualizadas")
    print("  - SÍ: En formulario y XML/XSD")
    print("  - Estructural: Generado automáticamente")
    print("  - Pendiente: No implementado")

if __name__ == '__main__':
    main()

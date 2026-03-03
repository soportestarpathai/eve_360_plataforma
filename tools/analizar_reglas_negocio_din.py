#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analiza Reglas de Negocio (col H) del instructivo DIN vs validación del formulario.
Agrega columna REGLAS NEGOCIO CUMPLE al Excel.
"""
import re
import sys
import os

# Validación del formulario DIN por tag (extraído de operacion_din.php)
# Formato: maxlength, pattern, required, type (number/date/select), step, catalog
FORM_VALIDATIONS = {
    'mes_reportado': {'maxlength': 6, 'pattern': r'\d{6}', 'required': True, 'type': 'text'},
    'clave_sujeto_obligado': {'maxlength': 13, 'required': True},
    'clave_actividad': {'maxlength': 3, 'required': True},
    'clave_entidad_colegiada': {'maxlength': 12},
    'exento': {'required': True, 'catalog': True},
    'referencia_aviso': {'maxlength': 14, 'required': True},
    'prioridad': {'required': True, 'catalog': True},
    'tipo_alerta': {'required': True, 'catalog': True},
    'descripcion_alerta': {'maxlength': 3000},
    'folio_modificacion': {'maxlength': 14},
    'descripcion_modificacion': {'maxlength': 3000},
    'tipo_operacion': {'required': True, 'catalog': True},
    'objeto_aviso_anterior': {'required': True, 'values': ['SI', 'NO']},
    'modificacion': {'required': True, 'values': ['SI', 'NO']},
    'entidad_federativa': {'required': True, 'catalog': True},
    'registro_licencia': {'maxlength': 200, 'required': True},
    'codigo_postal': {'maxlength': 5, 'pattern': r'\d{5}', 'required': True},
    'colonia': {'maxlength': 50, 'required': True},
    'calle': {'maxlength': 100, 'required': True},
    'tipo_desarrollo': {'required': True, 'catalog': True},
    'descripcion_desarrollo': {'maxlength': 3000},
    'monto_desarrollo': {'type': 'number', 'step': 0.01, 'required': True},
    'unidades_comercializadas': {'type': 'number', 'step': 0.01, 'required': True},
    'costo_unidad': {'type': 'number', 'step': 0.01, 'required': True},
    'otras_empresas': {'required': True, 'values': ['SI', 'NO']},
    'fecha_aportacion': {'type': 'date', 'required': True},
    'instrumento_monetario': {'catalog': True, 'required': True},
    'moneda': {'catalog': True, 'required': True},
    'monto_aportacion': {'type': 'number', 'step': 0.01, 'required': True},
    'aportacion_fideicomiso': {'required': True, 'values': ['SI', 'NO']},
    'nombre_institucion': {'maxlength': 254},
    'descripcion_bien': {'maxlength': 3000},
    'monto_estimado': {'type': 'number', 'step': 0.01},
    'numero_socios': {'type': 'number', 'min': 1, 'max': 99999999},
    'numero_terceros': {'type': 'number', 'min': 1, 'max': 99999999},
    'aportacion_anterior_socio': {'values': ['SI', 'NO']},
    'rfc_socio': {'maxlength': 13},
    'nombre': {'maxlength': 200},
    'apellido_paterno': {'maxlength': 200},
    'apellido_materno': {'maxlength': 200},
    'fecha_nacimiento': {'type': 'date'},
    'rfc': {'maxlength': 13},  # persona_fisica, representante
    'curp': {'maxlength': 18},
    'pais_nacionalidad': {'catalog': True},
    'actividad_economica': {'maxlength': 7, 'pattern': r'\d{7}'},
    'denominacion_razon': {'maxlength': 254},
    'fecha_constitucion': {'type': 'date'},
    'giro_mercantil': {'maxlength': 7, 'pattern': r'\d{7}'},
    'identificador_fideicomiso': {'maxlength': 40},
    'numero_exterior': {'maxlength': 56},
    'numero_interior': {'maxlength': 40},
    'estado_provincia': {'maxlength': 100},
    'ciudad_poblacion': {'maxlength': 100},
    'numero_telefono': {'maxlength': 12, 'pattern': r'\d{10,12}'},
    'correo_electronico': {'maxlength': 60},
    'clave_pais': {'maxlength': 4},
    'tipo_tercero': {'catalog': True},
    'descripcion_tercero': {'maxlength': 3000},
    'valor_inmueble_preventa': {'type': 'number', 'step': 0.01},
    'tipo_institucion': {'catalog': True},
    'institucion': {'maxlength': 254},
    'tipo_credito': {'catalog': True},
    'monto_prestamo': {'type': 'number', 'step': 0.01},
    'plazo_meses': {'type': 'number'},
    'fecha_emision': {'type': 'date'},
    'monto_solicitado': {'type': 'number', 'step': 0.01},
    'monto_recibido': {'type': 'number', 'step': 0.01},
}

# Tags soportados en formulario (para CUMPLE REGLAS)
TAGS_SOPORTADOS = {
    'archivo', 'informe', 'mes_reportado', 'sujeto_obligado', 'clave_entidad_colegiada',
    'clave_sujeto_obligado', 'clave_actividad', 'exento', 'aviso', 'referencia_aviso',
    'modificatorio', 'folio_modificacion', 'descripcion_modificacion', 'prioridad',
    'alerta', 'tipo_alerta', 'descripcion_alerta', 'detalle_operaciones', 'datos_operacion',
    'tipo_operacion', 'desarrollos_inmobiliarios', 'datos_desarrollo', 'objeto_aviso_anterior',
    'modificacion', 'entidad_federativa', 'registro_licencia', 'caracteristicas_desarrollo',
    'codigo_postal', 'colonia', 'calle', 'tipo_desarrollo', 'descripcion_desarrollo',
    'monto_desarrollo', 'unidades_comercializadas', 'costo_unidad', 'otras_empresas',
    'aportaciones', 'fecha_aportacion', 'tipo_aportacion', 'recursos_propios', 'datos_aportacion',
    'aportacion_numerario', 'instrumento_monetario', 'moneda', 'monto_aportacion',
    'aportacion_fideicomiso', 'nombre_institucion', 'aportacion_especie', 'descripcion_bien',
    'monto_estimado', 'socios', 'numero_socios', 'detalle_socios', 'datos_socio',
    'aportacion_anterior_socio', 'rfc_socio', 'tipo_persona_socio', 'persona_fisica',
    'nombre', 'apellido_paterno', 'apellido_materno', 'fecha_nacimiento', 'rfc', 'curp',
    'pais_nacionalidad', 'actividad_economica', 'persona_moral', 'denominacion_razon',
    'fecha_constitucion', 'giro_mercantil', 'representante_apoderado', 'fideicomiso',
    'identificador_fideicomiso', 'tipo_domicilio_socio', 'nacional', 'extranjero',
    'numero_exterior', 'numero_interior', 'pais', 'estado_provincia', 'ciudad_poblacion',
    'telefono', 'clave_pais', 'numero_telefono', 'correo_electronico', 'detalle_aportaciones',
    'terceros', 'numero_terceros', 'detalle_terceros', 'datos_tercero', 'tipo_tercero',
    'descripcion_tercero', 'tipo_persona_tercero', 'valor_inmueble_preventa',
    'prestamo_financiero', 'datos_prestamo', 'tipo_institucion', 'institucion', 'tipo_credito',
    'monto_prestamo', 'plazo_meses', 'prestamo_no_financiero', 'detalle_acreedores',
    'tipo_persona_acreedor', 'financiamiento_bursatil', 'fecha_emision', 'monto_solicitado', 'monto_recibido',
}

# Tags que son estructura/contenedor o sin validación específica en formulario
TAGS_SIN_VALIDACION = {
    'archivo', 'informe', 'aviso', 'sujeto_obligado', 'modificatorio', 'alerta',
    'detalle_operaciones', 'datos_operacion', 'desarrollos_inmobiliarios',
    'datos_desarrollo', 'caracteristicas_desarrollo', 'aportaciones',
    'tipo_aportacion', 'recursos_propios', 'datos_aportacion', 'aportacion_numerario',
    'aportacion_especie', 'socios', 'detalle_socios', 'datos_socio',
    'tipo_persona_socio', 'persona_fisica', 'persona_moral', 'fideicomiso',
    'representante_apoderado', 'tipo_domicilio_socio', 'nacional', 'extranjero',
    'telefono', 'terceros', 'detalle_terceros', 'datos_tercero',
    'tipo_persona_tercero', 'prestamo_financiero', 'datos_prestamo',
    'prestamo_no_financiero', 'detalle_acreedores', 'tipo_persona_acreedor',
    'financiamiento_bursatil', 'pais', 'detalle_aportaciones',
}

def extraer_tag(etiqueta_celda):
    if not etiqueta_celda:
        return None
    s = str(etiqueta_celda).strip()
    m = re.match(r'<?([a-z_]+)>?', s, re.I)
    return m.group(1).lower() if m else None

def parsear_reglas_negocio(texto, tag=None):
    """Extrae reglas clave del texto de reglas de negocio (col H)."""
    if not texto:
        return {}
    t = str(texto).upper()
    reglas = {}

    # Longitud: priorizar "máxima" sobre "mínima" para long_max
    m_max = re.search(r'M[ÁA]XIMA\s+(?:DE\s+)?(\d+)\s*(?:caracteres?|d[ií]gitos?)', t, re.I)
    m_min = re.search(r'M[IÍ]NIMA\s+(?:DE\s+)?(\d+)\s*(?:caracteres?|d[ií]gitos?)', t, re.I)
    m_exact = re.search(r'LONGITUD\s+ES\s+DE\s+(\d+)\s*(?:caracteres?|d[ií]gitos?)', t, re.I)
    if m_max:
        reglas['long_max'] = int(m_max.group(1))
    elif m_exact:
        reglas['long_max'] = int(m_exact.group(1))
    if m_min:
        reglas['long_min'] = int(m_min.group(1))
    # Evitar casos como "4 caracteres (1 entero...)" para montos - buscar "17" en contexto decimal
    if '17' in t and '14' in t and 'DECIMALES' in t:
        reglas['long_max'] = 17  # formato 14.2

    # Obligatorio
    if 'OBLIGATORIO' in t or 'ES OBLIGATORIO' in t:
        reglas['obligatorio'] = True
    if 'OPCIONAL' in t and 'NO OPCIONAL' not in t:
        reglas['obligatorio'] = False

    # Formato
    if 'AAAAMM' in t or 'AÑO-MES' in t:
        reglas['formato'] = 'AAAAMM'
    elif 'AAAAMMDD' in t:
        reglas['formato'] = 'AAAAMMDD'
    elif 'SI" O "NO' in t or '"SI" O "NO' in t or 'SI O NO' in t:
        reglas['valores'] = ['SI', 'NO']
    elif 'NUM[EÉ]RICO' in t and 'DECIMALES' in t:
        reglas['tipo'] = 'number_decimal'
    elif 'ENTERO NUM[EÉ]RICO' in t:
        reglas['tipo'] = 'integer'
    elif 'CAT[ÁA]LOGO' in t or 'CATALOGO' in t:
        reglas['catalogo'] = True

    # Patrón específico (solo cuando el contexto es claro)
    if tag == 'codigo_postal' and '5' in t and '0-9' in t:
        reglas['pattern'] = r'\d{5}'
    elif '6 D[IÍ]GITOS' in t and 'AAAAMM' in t:
        reglas['pattern'] = r'\d{6}'
    elif '18 CARACTERES' in t and ('CURP' in t or 'LLLLAAMMDD' in t):
        reglas['long_max'] = 18

    return reglas

def validar_reglas_negocio(tag, reglas_texto, form_valid):
    """
    Compara reglas de negocio parseadas con validación del formulario.
    Retorna (cumple: SÍ/Parcial/NO, observacion: str)
    """
    if tag in TAGS_SIN_VALIDACION:
        return 'SÍ', 'Estructura/contenedor, validación en campos hijos'

    reglas = parsear_reglas_negocio(reglas_texto, tag)
    if not reglas and not form_valid:
        return 'SÍ', 'Sin reglas específicas'
    if not form_valid:
        return 'NO', 'Campo no tiene validación en formulario'

    obs = []
    cumple = 'SÍ'

    # Comparar longitud (solo para campos texto, no numéricos)
    if 'long_max' in reglas and form_valid.get('type') != 'number':
        form_max = form_valid.get('maxlength')
        if form_max is not None:
            # clave_sujeto_obligado: instructivo 12-13, form 13 ✓
            if reglas.get('long_min') and reglas['long_max'] in (12, 13) and form_max == 13:
                pass  # OK
            elif reglas['long_max'] != form_max and form_max < reglas['long_max']:
                obs.append(f"Form maxlength {form_max} < instructivo {reglas['long_max']}")
                cumple = 'Parcial' if cumple == 'SÍ' else cumple

    # Comparar tipo numérico
    if reglas.get('tipo') == 'number_decimal':
        if form_valid.get('type') != 'number' and 'step' not in form_valid:
            obs.append('Esperado: numérico decimal')
            cumple = 'Parcial' if cumple == 'SÍ' else cumple

    # Comparar patrón
    if 'pattern' in reglas:
        if 'pattern' not in form_valid:
            obs.append(f"Falta validar patrón ({reglas['pattern']})")
            cumple = 'Parcial' if cumple == 'SÍ' else cumple

    # Catálogo
    if reglas.get('catalogo') and not form_valid.get('catalog'):
        obs.append('Debería usar catálogo UIF')
        cumple = 'Parcial' if cumple == 'SÍ' else cumple

    # Obligatorio
    if reglas.get('obligatorio') and not form_valid.get('required'):
        obs.append('Falta required en formulario')
        cumple = 'Parcial' if cumple == 'SÍ' else cumple

    return cumple, '; '.join(obs) if obs else ''

def main():
    try:
        import openpyxl
    except ImportError:
        print("Instalando openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        import openpyxl

    ruta_entrada = r'c:\Users\oscar\Downloads\instructivo_din(1).xlsx'
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ruta_salida = os.path.join(base_dir, 'instructivo_din_analizado.xlsx')
    ruta_downloads = r'c:\Users\oscar\Downloads\instructivo_din_analizado.xlsx'
    ruta_alt = os.path.join(base_dir, 'instructivo_din_completo.xlsx')  # alternativo si el principal está abierto

    if not os.path.exists(ruta_entrada):
        print(f"Error: No se encontró {ruta_entrada}")
        sys.exit(1)

    wb = openpyxl.load_workbook(ruta_entrada, data_only=False)
    ws = wb['DIN']

    # Columnas: A=1 NO, B=2 NOMBRE, C=3 ETIQUETA, D=4 OBLIG, E=5 TIPO, F=6 LONG, G=7 FORMATO, H=8 REGLAS, I=9 CATALOGOS
    col_etiqueta = 3
    col_reglas = 8

    # Insertar 4 columnas: CUMPLE REGLAS, OBSERVACIONES, REGLAS NEGOCIO CUMPLE, OBS. REGLAS
    headers_row1 = [str(ws.cell(1, c).value or '').upper() for c in range(1, min(15, ws.max_column + 5))]
    if 'REGLAS NEGOCIO' in ' '.join(headers_row1):
        # Ya tiene columnas, buscar índices
        col_reglas_cumple = col_obs_reglas = None
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(1, c).value or '').upper()
            if 'REGLAS NEGOCIO' in h and 'OBS' not in h:
                col_reglas_cumple = c
            elif col_reglas_cumple and c == col_reglas_cumple + 1:
                col_obs_reglas = c
                break
        if col_reglas_cumple is None:
            ws.cell(1, ws.max_column + 1, 'REGLAS NEGOCIO CUMPLE')
            ws.cell(1, ws.max_column + 2, 'OBS. REGLAS')
            col_reglas_cumple = ws.max_column - 1
            col_obs_reglas = ws.max_column
    else:
        # Archivo original: insertar todas las columnas
        ws.insert_cols(10, 4)
        ws.cell(1, 10, 'CUMPLE REGLAS')
        ws.cell(1, 11, 'OBSERVACIONES')
        ws.cell(1, 12, 'REGLAS NEGOCIO CUMPLE')
        ws.cell(1, 13, 'OBS. REGLAS')
        col_reglas_cumple = 12
        col_obs_reglas = 13

    col_cumple = 10
    col_obs = 11

    for row_idx in range(2, ws.max_row + 1):
        etiqueta_val = ws.cell(row_idx, col_etiqueta).value
        reglas_val = ws.cell(row_idx, col_reglas).value
        tag = extraer_tag(etiqueta_val)

        # CUMPLE REGLAS (campo en formulario)
        if tag and (tag in TAGS_SOPORTADOS or tag in TAGS_SIN_VALIDACION):
            cumple_campo, obs_campo = 'SÍ', ''
        elif tag:
            cumple_campo, obs_campo = 'NO', 'Campo no implementado'
        else:
            cumple_campo, obs_campo = 'NO', 'Tag no identificado'

        # REGLAS NEGOCIO (validación acorde al instructivo)
        form_valid = FORM_VALIDATIONS.get(tag, {}) if tag else {}
        cumple_reglas, obs_reglas = validar_reglas_negocio(tag, reglas_val, form_valid)

        ws.cell(row_idx, col_cumple, cumple_campo)
        ws.cell(row_idx, col_obs, obs_campo)
        ws.cell(row_idx, col_reglas_cumple, cumple_reglas)
        ws.cell(row_idx, col_obs_reglas, obs_reglas or '')

    guardado = []
    for path in [ruta_salida, ruta_downloads, ruta_alt]:
        try:
            wb.save(path)
            guardado.append(path)
        except PermissionError:
            continue
    if guardado:
        print("Archivo guardado en:", ", ".join(guardado))
    else:
        print("Error: No se pudo guardar (cierre Excel si está abierto). Intentando ruta alterna...")
        wb.save(ruta_alt)
        print("Guardado en:", ruta_alt)
    print("Columnas: CUMPLE REGLAS, OBSERVACIONES, REGLAS NEGOCIO CUMPLE, OBS. REGLAS")

if __name__ == '__main__':
    main()

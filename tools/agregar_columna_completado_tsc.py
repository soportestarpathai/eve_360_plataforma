#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agrega columna COMPLETADO al final de instructivo_tsc.xlsx
Ejecutar: python tools/agregar_columna_completado_tsc.py
"""
import os
import sys

def main():
    try:
        import openpyxl
    except ImportError:
        print("Instalando openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        import openpyxl

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    # Buscar instructivo_tsc.xlsx en varias ubicaciones
    posibles_rutas = [
        os.path.join(base_dir, 'instructivo_tsc.xlsx'),
        os.path.join(base_dir, 'docs', 'instructivo_tsc.xlsx'),
        os.path.join(os.path.expanduser('~'), 'Downloads', 'instructivo_tsc.xlsx'),
        os.path.join(os.path.expanduser('~'), 'Descargas', 'instructivo_tsc.xlsx'),
    ]
    
    ruta_entrada = None
    for r in posibles_rutas:
        if os.path.exists(r):
            ruta_entrada = r
            break
    
    if ruta_entrada is None:
        print("No se encontró instructivo_tsc.xlsx")
        print("Colócalo en el proyecto (raíz o docs/) o en Descargas")
        print("Rutas buscadas:")
        for r in posibles_rutas:
            print(f"  - {r}")
        sys.exit(1)

    print(f"Procesando: {ruta_entrada}")

    wb = openpyxl.load_workbook(ruta_entrada, data_only=False)
    # Usar la primera hoja o la que se llame TSC
    ws_name = 'TSC' if 'TSC' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]

    # Obtener número de columnas actual (la última usada)
    max_col = ws.max_column
    
    # Agregar columna COMPLETADO al final
    col_completado = max_col + 1
    ws.cell(1, col_completado, 'COMPLETADO')
    
    # Dejar celdas vacías para que el usuario marque cuando incorpore cada campo
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, col_completado, '')  # Vacío = pendiente
    
    wb.save(ruta_entrada)
    print(f"✓ Columna 'COMPLETADO' agregada al final (columna {col_completado})")
    print(f"✓ Archivo guardado: {ruta_entrada}")

if __name__ == '__main__':
    main()
